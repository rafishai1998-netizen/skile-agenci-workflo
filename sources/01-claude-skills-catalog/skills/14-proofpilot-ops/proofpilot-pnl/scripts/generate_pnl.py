#!/usr/bin/env python3
"""
ProofPilot P&L Generator

Usage:
    python generate_pnl.py --month 12 --year 2025 --output /mnt/user-data/outputs/

Requires data dict with:
    - revenue: dict of revenue line items
    - cogs: dict of COGS line items  
    - opex: dict of OpEx line items
    - owner_pay: dict of owner compensation
    - prior_month: optional dict with same structure for comparison
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ProofPilot Brand Colors
COLORS = {
    'navy': '00184D',
    'blue': '0051FF', 
    'light_blue': 'E6F0FF',
    'white': 'FFFFFF',
    'gray': '4A5568',
    'light_gray': 'F7FAFC',
    'green': '10B981',
    'orange': 'F59E0B',
    'red': 'EF4444',
    'purple': '7C3AED',
    'tax_red': 'DC2626',
}

def create_pnl_workbook(data: dict, month: int, year: int, prior_data: dict = None) -> Workbook:
    """Create P&L workbook from data dict."""
    
    wb = Workbook()
    ws = wb.active
    
    month_name = datetime(year, month, 1).strftime('%B')
    ws.title = f"{month_name} {year} P&L"
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 30
    
    # Styles
    header_font = Font(name='Arial', bold=True, size=12, color=COLORS['white'])
    section_font = Font(name='Arial', bold=True, size=11, color=COLORS['white'])
    normal_font = Font(name='Arial', size=11)
    title_font = Font(name='Arial', bold=True, size=18, color=COLORS['navy'])
    
    header_fill = PatternFill('solid', fgColor=COLORS['navy'])
    section_fill = PatternFill('solid', fgColor=COLORS['blue'])
    alt_fill = PatternFill('solid', fgColor=COLORS['light_gray'])
    
    thin_border = Border(
        left=Side(style='thin', color=COLORS['gray']),
        right=Side(style='thin', color=COLORS['gray']),
        top=Side(style='thin', color=COLORS['gray']),
        bottom=Side(style='thin', color=COLORS['gray'])
    )
    
    # Title
    ws['A1'] = 'PROOFPILOT P&L'
    ws['A1'].font = title_font
    ws['A2'] = f'{month_name} {year}'
    ws['A2'].font = Font(name='Arial', italic=True, size=11, color=COLORS['gray'])
    
    # Headers
    row = 4
    headers = ['Category', month_name, 'Prior Month', 'Change', '% Rev', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 6
    
    def add_section(title: str, items: dict, prior_items: dict, total_rev: float, 
                    section_color: str, total_color: str, total_label: str):
        nonlocal row
        
        # Section header
        ws.cell(row=row, column=1, value=title).font = section_font
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=COLORS[section_color])
            ws.cell(row=row, column=c).border = thin_border
        row += 1
        
        # Items
        total_current = 0
        total_prior = 0
        for i, (name, values) in enumerate(items.items()):
            current = values.get('amount', 0)
            prior = prior_items.get(name, {}).get('amount', 0) if prior_items else 0
            note = values.get('note', '')
            
            total_current += current
            total_prior += prior
            
            ws.cell(row=row, column=1, value=name).font = normal_font
            ws.cell(row=row, column=2, value=current).number_format = '"$"#,##0'
            ws.cell(row=row, column=3, value=prior).number_format = '"$"#,##0'
            ws.cell(row=row, column=4, value=current - prior).number_format = '"$"#,##0'
            if total_rev > 0:
                ws.cell(row=row, column=5, value=current / total_rev).number_format = '0.0%'
            ws.cell(row=row, column=6, value=note).font = normal_font
            
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border
                if i % 2 == 1:
                    ws.cell(row=row, column=c).fill = alt_fill
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value=total_label).font = Font(name='Arial', bold=True, size=11, color=COLORS['white'])
        ws.cell(row=row, column=2, value=total_current).number_format = '"$"#,##0'
        ws.cell(row=row, column=3, value=total_prior).number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=total_current - total_prior).number_format = '"$"#,##0'
        if total_rev > 0:
            ws.cell(row=row, column=5, value=total_current / total_rev).number_format = '0.0%'
        
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=COLORS[total_color])
            ws.cell(row=row, column=c).font = Font(name='Arial', bold=True, size=11, color=COLORS['white'])
            ws.cell(row=row, column=c).border = thin_border
        
        row += 2
        return total_current, total_prior
    
    def add_calculated_row(label: str, current: float, prior: float, total_rev: float,
                           color: str, note: str = ''):
        nonlocal row
        
        ws.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=12, color=COLORS['white'])
        ws.cell(row=row, column=2, value=current).number_format = '"$"#,##0'
        ws.cell(row=row, column=3, value=prior).number_format = '"$"#,##0'
        ws.cell(row=row, column=4, value=current - prior).number_format = '"$"#,##0'
        if total_rev > 0:
            ws.cell(row=row, column=5, value=current / total_rev).number_format = '0.0%'
        ws.cell(row=row, column=6, value=note).font = Font(name='Arial', size=10, color=COLORS['white'])
        
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=COLORS[color])
            ws.cell(row=row, column=c).font = Font(name='Arial', bold=True, size=12, color=COLORS['white'])
            ws.cell(row=row, column=c).border = thin_border
        
        row += 2
    
    # Get prior month data
    prior_rev = prior_data.get('revenue', {}) if prior_data else {}
    prior_cogs = prior_data.get('cogs', {}) if prior_data else {}
    prior_opex = prior_data.get('opex', {}) if prior_data else {}
    prior_owner = prior_data.get('owner_pay', {}) if prior_data else {}
    
    # Revenue
    total_rev, prior_total_rev = add_section(
        'REVENUE', data['revenue'], prior_rev, 
        sum(v.get('amount', 0) for v in data['revenue'].values()),
        'blue', 'green', 'TOTAL REVENUE'
    )
    
    # COGS
    total_cogs, prior_total_cogs = add_section(
        'COST OF GOODS SOLD', data['cogs'], prior_cogs,
        total_rev, 'blue', 'red', 'TOTAL COGS'
    )
    
    # Gross Profit
    gp = total_rev - total_cogs
    prior_gp = prior_total_rev - prior_total_cogs
    add_calculated_row('GROSS PROFIT', gp, prior_gp, total_rev, 'green', 'Gross Margin')
    
    # OpEx
    total_opex, prior_total_opex = add_section(
        'OPERATING EXPENSES', data['opex'], prior_opex,
        total_rev, 'blue', 'orange', 'TOTAL OPEX'
    )
    
    # EBITDA
    ebitda = gp - total_opex
    prior_ebitda = prior_gp - prior_total_opex
    add_calculated_row('EBITDA', ebitda, prior_ebitda, total_rev, 'blue', 'Before taxes & owner pay')
    
    # Tax Reserve
    tax_items = {
        'Self-Employment (15.3%)': {'amount': ebitda * 0.153},
        'Federal Income (12%)': {'amount': ebitda * 0.12},
        'Arizona State (2.5%)': {'amount': ebitda * 0.025},
    }
    prior_tax_items = {
        'Self-Employment (15.3%)': {'amount': prior_ebitda * 0.153},
        'Federal Income (12%)': {'amount': prior_ebitda * 0.12},
        'Arizona State (2.5%)': {'amount': prior_ebitda * 0.025},
    }
    total_tax, prior_total_tax = add_section(
        'TAX RESERVE (29.8%)', tax_items, prior_tax_items,
        total_rev, 'tax_red', 'tax_red', 'TOTAL TAX RESERVE'
    )
    
    # After Tax Income
    after_tax = ebitda - total_tax
    prior_after_tax = prior_ebitda - prior_total_tax
    add_calculated_row('INCOME AFTER TAX', after_tax, prior_after_tax, total_rev, 'green', 'Available for owner')
    
    # Owner Compensation
    total_owner, prior_total_owner = add_section(
        'OWNER COMPENSATION', data['owner_pay'], prior_owner,
        total_rev, 'blue', 'purple', 'TOTAL OWNER PAY'
    )
    
    # Net Profit
    net = after_tax - total_owner
    prior_net = prior_after_tax - prior_total_owner
    add_calculated_row('NET PROFIT', net, prior_net, total_rev, 'navy', 'After taxes & owner pay')
    
    return wb


def save_pnl(wb: Workbook, month: int, year: int, output_dir: str = '/mnt/user-data/outputs/'):
    """Save workbook to output directory."""
    month_name = datetime(year, month, 1).strftime('%B').lower()
    filename = f'{output_dir}proofpilot_pnl_{month_name}_{year}.xlsx'
    wb.save(filename)
    return filename


# Example usage
if __name__ == '__main__':
    # Example data structure
    example_data = {
        'revenue': {
            'SEO Retainers': {'amount': 21513.69, 'note': ''},
            'Ads Management (Meta)': {'amount': 1000, 'note': ''},
            'Funnels': {'amount': 1000, 'note': 'New service'},
        },
        'cogs': {
            'Jo Paula': {'amount': 2500, 'note': 'White label'},
            'Marcos': {'amount': 2600, 'note': '$2K + $600 bonus'},
            'Muhaiminul': {'amount': 1191.40, 'note': 'SEO'},
        },
        'opex': {
            'Software - AI Tools': {'amount': 212.83, 'note': 'Claude, Genspark'},
            'Software - SEO Tools': {'amount': 226.54, 'note': ''},
        },
        'owner_pay': {
            'Owner Draw': {'amount': 11000, 'note': 'SoFi transfers'},
            'SEP-IRA': {'amount': 0, 'note': ''},
        }
    }
    
    wb = create_pnl_workbook(example_data, 12, 2025)
    filename = save_pnl(wb, 12, 2025)
    print(f'Saved to {filename}')
