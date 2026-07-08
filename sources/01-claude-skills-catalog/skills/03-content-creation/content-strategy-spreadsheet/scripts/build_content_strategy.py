#!/usr/bin/env python3
"""
Content Strategy Spreadsheet Builder
Generates professionally styled Excel spreadsheets matching the Trading Academy template.
"""

import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR CONSTANTS - Extracted from Trading Academy template
# =============================================================================

# Primary brand colors
PRIMARY_BLUE = "151F6D"
ACCENT_ORANGE = "EF6430"
WHITE = "FFFFFF"
BODY_TEXT = "4A4A4A"

# Background colors
LIGHT_GRAY = "F1F0F0"
SECTION_HEADER_BG = "E8EAF6"
AUTHORITY_BG = "FDE8DC"

# Stage colors (funnel)
STAGE_AWARENESS = "E8F5E9"
STAGE_CONSIDERATION = "FFF8E1"
STAGE_DECISION = "FBE9E7"

# Score colors (psychographic)
SCORE_HIGH = "C8E6C9"
SCORE_MED = "FFF9C4"
SCORE_LOW = "FFCDD2"

# Priority colors (roadmap)
PRIORITY_P1 = "FFCDD2"
PRIORITY_P2 = "FFF9C4"
PRIORITY_P3 = "C8E6C9"
PRIORITY_P4 = "BBDEFB"

# Content type badge colors
TYPE_COLORS = {
    "Pillar": "1565C0",
    "Research": "6A1B9A",
    "Supporting": "2E7D32",
    "Comparison": "F57C00",
    "How-To": "00838F",
    "Tutorial": "5D4037",
    "Assessment": "1565C0",
    "Calculator": "2E7D32",
    "Simulator": "F57C00",
    "Quiz": "6A1B9A",
    "Planner": "00838F",
    "E-book": "1565C0",
    "Guide": "2E7D32",
    "Checklist": "F57C00",
    "Template": "6A1B9A",
    "Worksheet": "00838F",
}

# =============================================================================
# STYLE HELPERS
# =============================================================================

def make_fill(hex_color):
    """Create a solid fill from hex color."""
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_font(color=BODY_TEXT, bold=False, size=12):
    """Create a font with specified properties."""
    return Font(name="Arial", color=color, bold=bold, size=size)

def make_border(style="thin"):
    """Create a border with specified style."""
    side = Side(style=style, color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """Set cell value and styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

def get_stage_fill(stage):
    """Get fill color for funnel stage."""
    stage_lower = stage.lower() if stage else ""
    if "awareness" in stage_lower or "top" in stage_lower:
        return make_fill(STAGE_AWARENESS)
    elif "consideration" in stage_lower or "middle" in stage_lower:
        return make_fill(STAGE_CONSIDERATION)
    elif "decision" in stage_lower or "bottom" in stage_lower:
        return make_fill(STAGE_DECISION)
    return None

def get_score_fill(score):
    """Get fill color for psychographic score."""
    if isinstance(score, (int, float)):
        if score >= 4:
            return make_fill(SCORE_HIGH)
        elif score >= 3:
            return make_fill(SCORE_MED)
        else:
            return make_fill(SCORE_LOW)
    return None

def get_priority_fill(priority):
    """Get fill color for priority badge."""
    p = str(priority).upper()
    if p in ("P1", "1"):
        return make_fill(PRIORITY_P1)
    elif p in ("P2", "2"):
        return make_fill(PRIORITY_P2)
    elif p in ("P3", "3"):
        return make_fill(PRIORITY_P3)
    elif p in ("P4", "4"):
        return make_fill(PRIORITY_P4)
    return None

def get_type_fill(content_type):
    """Get fill color for content type badge."""
    color = TYPE_COLORS.get(content_type, "757575")
    return make_fill(color)

# =============================================================================
# SHEET BUILDERS
# =============================================================================

def build_strategy_overview(wb, data):
    """Build the Strategy Overview sheet."""
    ws = wb.create_sheet("Strategy Overview")
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 28
    
    row = 1
    
    # Title
    set_cell(ws, row, 1, data.get("brand_name", "").upper(),
             font=make_font(PRIMARY_BLUE, bold=True, size=13))
    row += 1
    
    # Subtitle
    set_cell(ws, row, 1, data.get("subtitle", ""),
             font=make_font(ACCENT_ORANGE, bold=True, size=13))
    row += 2
    
    # THE OPPORTUNITY section
    set_cell(ws, row, 1, "THE OPPORTUNITY",
             font=make_font(PRIMARY_BLUE, bold=True, size=13),
             fill=make_fill(SECTION_HEADER_BG))
    for col in range(2, 5):
        set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
    row += 1
    
    # Opportunity text
    opp_cell = set_cell(ws, row, 1, data.get("opportunity", ""),
                        font=make_font(BODY_TEXT, size=12),
                        alignment=Alignment(wrap_text=True, vertical="top"))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 60
    row += 3
    
    # FUNNEL STRUCTURE section
    set_cell(ws, row, 1, "FUNNEL STRUCTURE",
             font=make_font(PRIMARY_BLUE, bold=True, size=13),
             fill=make_fill(SECTION_HEADER_BG))
    for col in range(2, 5):
        set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
    row += 1
    
    # Funnel table headers
    headers = ["Stage", "Goal", "Content Focus", "Conversion Point"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, row, col, header,
                 font=make_font(WHITE, bold=True, size=12),
                 fill=make_fill(PRIMARY_BLUE),
                 border=make_border())
    row += 1
    
    # Funnel rows
    for item in data.get("funnel", []):
        stage = item.get("stage", "")
        stage_fill = get_stage_fill(stage)
        
        set_cell(ws, row, 1, stage.upper() if stage else "",
                 font=make_font(BODY_TEXT, bold=True, size=12),
                 fill=stage_fill,
                 border=make_border())
        set_cell(ws, row, 2, item.get("goal", ""),
                 font=make_font(BODY_TEXT, size=12),
                 border=make_border())
        set_cell(ws, row, 3, item.get("content", ""),
                 font=make_font(BODY_TEXT, size=12),
                 border=make_border(),
                 alignment=Alignment(wrap_text=True))
        set_cell(ws, row, 4, item.get("cta", ""),
                 font=make_font(BODY_TEXT, size=12),
                 border=make_border())
        row += 1
    
    row += 1
    
    # TARGET AUDIENCES section
    set_cell(ws, row, 1, "TARGET AUDIENCES",
             font=make_font(PRIMARY_BLUE, bold=True, size=13),
             fill=make_fill(SECTION_HEADER_BG))
    for col in range(2, 5):
        set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
    row += 1
    
    # Audience table headers
    headers = ["Profile", "%", "Description", "Entry Stage"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, row, col, header,
                 font=make_font(WHITE, bold=True, size=12),
                 fill=make_fill(PRIMARY_BLUE),
                 border=make_border())
    row += 1
    
    # Audience rows (alternating colors)
    for i, item in enumerate(data.get("audiences", [])):
        row_fill = make_fill(LIGHT_GRAY) if i % 2 == 0 else make_fill(WHITE)
        
        set_cell(ws, row, 1, item.get("name", ""),
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border())
        set_cell(ws, row, 2, item.get("percentage", ""),
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border())
        desc = item.get("description", "") or item.get("content_needs", "")
        set_cell(ws, row, 3, desc,
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border(),
                 alignment=Alignment(wrap_text=True))
        entry = item.get("entry_stage", "") or item.get("content_needs", "")[:20] if item.get("content_needs") else ""
        set_cell(ws, row, 4, entry,
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border())
        row += 1
    
    row += 1
    
    # CONTENT PILLARS section
    set_cell(ws, row, 1, "CONTENT PILLARS",
             font=make_font(PRIMARY_BLUE, bold=True, size=13),
             fill=make_fill(SECTION_HEADER_BG))
    for col in range(2, 4):
        set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
    row += 1
    
    # Pillar table headers
    headers = ["Pillar", "Course Alignment", "Conversion Path"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, row, col, header,
                 font=make_font(WHITE, bold=True, size=12),
                 fill=make_fill(PRIMARY_BLUE),
                 border=make_border())
    row += 1
    
    # Pillar rows (alternating colors)
    for i, item in enumerate(data.get("pillars", [])):
        row_fill = make_fill(LIGHT_GRAY) if i % 2 == 0 else make_fill(WHITE)
        
        name = item.get("name", "") or item.get("topics", "")[:30]
        set_cell(ws, row, 1, name,
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border())
        alignment_val = item.get("course_alignment", "") or item.get("product", "")
        set_cell(ws, row, 2, alignment_val,
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border())
        conversion = item.get("conversion_path", "") or item.get("conversion", "")
        set_cell(ws, row, 3, conversion,
                 font=make_font(BODY_TEXT, size=12),
                 fill=row_fill,
                 border=make_border())
        row += 1
    
    return ws


def build_customer_profiles(wb, data):
    """Build the Customer Profiles sheet."""
    ws = wb.create_sheet("Customer Profiles")
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    
    row = 1
    
    for profile in data.get("profiles", []):
        # Profile header
        name = profile.get("name", "Profile")
        pct = profile.get("percentage", "")
        header_text = f"{name} ({pct})" if pct else name
        
        set_cell(ws, row, 1, header_text,
                 font=make_font(PRIMARY_BLUE, bold=True, size=13),
                 fill=make_fill(SECTION_HEADER_BG))
        for col in range(2, 5):
            set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
        row += 1
        
        # Description
        set_cell(ws, row, 1, profile.get("description", ""),
                 font=make_font(BODY_TEXT, size=12),
                 alignment=Alignment(wrap_text=True, vertical="top"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 45
        row += 2
        
        # Demographics
        set_cell(ws, row, 1, "Demographics:",
                 font=make_font(PRIMARY_BLUE, bold=True, size=12))
        set_cell(ws, row, 2, profile.get("demographics", ""),
                 font=make_font(BODY_TEXT, size=12))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 2
        
        # Pain Points
        set_cell(ws, row, 1, "Pain Points:",
                 font=make_font(PRIMARY_BLUE, bold=True, size=12))
        row += 1
        for pain in profile.get("pain_points", []):
            set_cell(ws, row, 1, f"• {pain}",
                     font=make_font(BODY_TEXT, size=12),
                     alignment=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
        row += 1
        
        # Motivations
        set_cell(ws, row, 1, "Motivations:",
                 font=make_font(PRIMARY_BLUE, bold=True, size=12))
        row += 1
        for mot in profile.get("motivations", []):
            set_cell(ws, row, 1, f"• {mot}",
                     font=make_font(BODY_TEXT, size=12),
                     alignment=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
        row += 1
        
        # Objections
        set_cell(ws, row, 1, "Objections:",
                 font=make_font(PRIMARY_BLUE, bold=True, size=12))
        row += 1
        for obj in profile.get("objections", []):
            set_cell(ws, row, 1, f"• {obj}",
                     font=make_font(BODY_TEXT, size=12),
                     alignment=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
        row += 1
        
        # Psychographic Factors table
        set_cell(ws, row, 1, "Psychographic Factors:",
                 font=make_font(PRIMARY_BLUE, bold=True, size=12))
        row += 1
        
        # Table headers
        psych_headers = ["Factor", "Score", "Level", "Notes"]
        for col, header in enumerate(psych_headers, 1):
            set_cell(ws, row, col, header,
                     font=make_font(WHITE, bold=True, size=11),
                     fill=make_fill(PRIMARY_BLUE),
                     border=make_border())
        row += 1
        
        # Psychographic rows
        for pf in profile.get("psychographic_factors", []):
            score = pf.get("score", 0)
            score_fill = get_score_fill(score)
            
            # Determine level from score
            if score >= 4:
                level = "High"
            elif score >= 3:
                level = "Medium"
            else:
                level = "Low"
            
            set_cell(ws, row, 1, pf.get("factor", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            set_cell(ws, row, 2, score,
                     font=make_font(BODY_TEXT, size=11),
                     fill=score_fill,
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            set_cell(ws, row, 3, level,
                     font=make_font(BODY_TEXT, size=11),
                     fill=score_fill,
                     border=make_border())
            set_cell(ws, row, 4, pf.get("notes", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            row += 1
        
        row += 1
        
        # Authority Hook
        set_cell(ws, row, 1, "Authority Hook:",
                 font=make_font(PRIMARY_BLUE, bold=True, size=12))
        row += 1
        set_cell(ws, row, 1, profile.get("authority_hook", ""),
                 font=make_font(ACCENT_ORANGE, bold=True, size=12),
                 fill=make_fill(AUTHORITY_BG),
                 alignment=Alignment(wrap_text=True, vertical="top"))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 60
        row += 3
    
    return ws


def build_content_hub(wb, data):
    """Build the Content Hub sheet."""
    ws = wb.create_sheet("Content Hub")
    
    # Column widths
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    
    row = 1
    
    for section in data.get("content_sections", []):
        # Section header with topical volume
        section_name = section.get("name", "Content Section")
        section_vol = section.get("volume", "")
        header_text = f"{section_name} | Topical Volume: {section_vol}" if section_vol else section_name
        
        set_cell(ws, row, 1, header_text,
                 font=make_font(PRIMARY_BLUE, bold=True, size=13),
                 fill=make_fill(SECTION_HEADER_BG))
        for col in range(2, 9):
            set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
        row += 1
        
        # Description
        if section.get("description"):
            set_cell(ws, row, 1, section.get("description", ""),
                     font=make_font(BODY_TEXT, size=11),
                     alignment=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            row += 1
        row += 1
        
        # Table headers
        headers = ["Type", "☐", "Title", "URL", "Primary Keyword", "Secondary Keywords", "Vol.", "Stage"]
        for col, header in enumerate(headers, 1):
            set_cell(ws, row, col, header,
                     font=make_font(WHITE, bold=True, size=11),
                     fill=make_fill(PRIMARY_BLUE),
                     border=make_border())
        row += 1
        
        # Content items
        for item in section.get("items", []):
            content_type = item.get("type", "")
            stage = item.get("stage", "")
            stage_fill = get_stage_fill(stage)
            type_fill = get_type_fill(content_type)
            
            # Type badge
            set_cell(ws, row, 1, content_type,
                     font=make_font(WHITE, bold=True, size=11),
                     fill=type_fill,
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Checkbox
            set_cell(ws, row, 2, "☐",
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Title
            set_cell(ws, row, 3, item.get("title", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            # URL
            set_cell(ws, row, 4, item.get("url", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Primary Keyword
            set_cell(ws, row, 5, item.get("primary_keyword", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Secondary Keywords
            set_cell(ws, row, 6, item.get("secondary_keywords", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            # Volume
            vol = item.get("topical_volume", "") or item.get("est_volume", "")
            set_cell(ws, row, 7, vol,
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Stage
            set_cell(ws, row, 8, stage,
                     font=make_font(BODY_TEXT, size=11),
                     fill=stage_fill,
                     border=make_border())
            
            row += 1
        
        row += 2
    
    return ws


def build_interactive_tools(wb, data):
    """Build the Interactive Tools sheet."""
    ws = wb.create_sheet("Interactive Tools")
    
    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 45
    
    row = 1
    
    # Header row
    headers = ["Type", "Name", "URL", "Primary Keyword", "Vol.", "Stage", "Target Profile", "Purpose"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, row, col, header,
                 font=make_font(WHITE, bold=True, size=11),
                 fill=make_fill(PRIMARY_BLUE),
                 border=make_border())
    row += 1
    
    # Group tools by type
    tools_by_type = {}
    for tool in data.get("tools", []):
        tool_type = tool.get("type", "Tool")
        if tool_type not in tools_by_type:
            tools_by_type[tool_type] = []
        tools_by_type[tool_type].append(tool)
    
    # Output each type group
    for tool_type, tools in tools_by_type.items():
        for tool in tools:
            stage = tool.get("stage", "")
            stage_fill = get_stage_fill(stage)
            type_fill = get_type_fill(tool_type)
            
            # Type badge
            set_cell(ws, row, 1, tool_type,
                     font=make_font(WHITE, bold=True, size=11),
                     fill=type_fill,
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Name
            set_cell(ws, row, 2, tool.get("name", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # URL
            set_cell(ws, row, 3, tool.get("url", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Primary Keyword
            set_cell(ws, row, 4, tool.get("primary_keyword", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Volume
            set_cell(ws, row, 5, tool.get("volume", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Stage
            set_cell(ws, row, 6, stage,
                     font=make_font(BODY_TEXT, size=11),
                     fill=stage_fill,
                     border=make_border())
            
            # Target Profile
            set_cell(ws, row, 7, tool.get("target_profile", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Purpose
            set_cell(ws, row, 8, tool.get("purpose", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            row += 1
        
        row += 1  # Blank row between types
    
    return ws


def build_ebooks_guides(wb, data):
    """Build the E-books & Guides sheet."""
    ws = wb.create_sheet("E-books & Guides")
    
    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 72
    ws.column_dimensions['C'].width = 56
    ws.column_dimensions['D'].width = 38
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 25
    
    row = 1
    
    # Header row
    headers = ["Type", "Title", "URL", "Related Articles", "Primary Keyword", "Vol.", "Target Profile", "Lead Magnet For"]
    for col, header in enumerate(headers, 1):
        set_cell(ws, row, col, header,
                 font=make_font(WHITE, bold=True, size=11),
                 fill=make_fill(PRIMARY_BLUE),
                 border=make_border())
    row += 1
    
    # Group ebooks by type
    ebooks_by_type = {}
    for ebook in data.get("ebooks", []):
        ebook_type = ebook.get("type", "E-book")
        if ebook_type not in ebooks_by_type:
            ebooks_by_type[ebook_type] = []
        ebooks_by_type[ebook_type].append(ebook)
    
    # Output each type group
    for ebook_type, ebooks in ebooks_by_type.items():
        for ebook in ebooks:
            type_fill = get_type_fill(ebook_type)
            
            # Type badge
            set_cell(ws, row, 1, ebook_type,
                     font=make_font(WHITE, bold=True, size=11),
                     fill=type_fill,
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Title
            set_cell(ws, row, 2, ebook.get("title", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            # URL
            set_cell(ws, row, 3, ebook.get("url", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Related Articles
            set_cell(ws, row, 4, ebook.get("related_articles", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            # Primary Keyword
            set_cell(ws, row, 5, ebook.get("primary_keyword", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Volume
            set_cell(ws, row, 6, ebook.get("volume", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Target Profile
            set_cell(ws, row, 7, ebook.get("target_profile", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Lead Magnet For
            set_cell(ws, row, 8, ebook.get("lead_magnet_for", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            row += 1
        
        row += 1  # Blank row between types
    
    return ws


def build_roadmap(wb, data):
    """Build the Roadmap sheet."""
    ws = wb.create_sheet("Roadmap")
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    
    row = 1
    
    # Title
    set_cell(ws, row, 1, "ROADMAP & PLAN OF ACTION",
             font=make_font(PRIMARY_BLUE, bold=True, size=13))
    row += 1
    
    # Subtitle
    subtitle = data.get("roadmap_subtitle", data.get("subtitle", ""))
    set_cell(ws, row, 1, subtitle,
             font=make_font(ACCENT_ORANGE, bold=True, size=13))
    row += 2
    
    # Priority Legend
    set_cell(ws, row, 1, "PRIORITY LEGEND",
             font=make_font(PRIMARY_BLUE, bold=True, size=12))
    row += 1
    
    legend = [
        ("P1", "Critical Path", "Must complete before other work can proceed"),
        ("P2", "High Priority", "Core content that drives conversions"),
        ("P3", "Medium Priority", "Supporting content and expansion"),
        ("P4", "Lower Priority", "Nice to have, build as resources allow"),
    ]
    
    for priority, label, desc in legend:
        set_cell(ws, row, 1, priority,
                 font=make_font(BODY_TEXT, bold=True, size=11),
                 fill=get_priority_fill(priority),
                 border=make_border(),
                 alignment=Alignment(horizontal="center"))
        set_cell(ws, row, 2, label,
                 font=make_font(BODY_TEXT, bold=True, size=11),
                 border=make_border())
        set_cell(ws, row, 3, desc,
                 font=make_font(BODY_TEXT, size=11),
                 border=make_border())
        row += 1
    
    row += 1
    
    # Phases
    for phase in data.get("phases", []):
        # Phase header
        phase_name = phase.get("name", "Phase")
        set_cell(ws, row, 1, phase_name.upper(),
                 font=make_font(PRIMARY_BLUE, bold=True, size=13),
                 fill=make_fill(SECTION_HEADER_BG))
        for col in range(2, 8):
            set_cell(ws, row, col, "", fill=make_fill(SECTION_HEADER_BG))
        row += 1
        
        # Phase description
        if phase.get("description"):
            set_cell(ws, row, 1, phase.get("description", ""),
                     font=make_font(BODY_TEXT, size=11),
                     alignment=Alignment(wrap_text=True))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1
        
        # Table headers
        headers = ["Priority", "Deliverable", "Type", "Why It Matters", "Dependencies", "Success Metric", "Owner"]
        for col, header in enumerate(headers, 1):
            set_cell(ws, row, col, header,
                     font=make_font(WHITE, bold=True, size=11),
                     fill=make_fill(PRIMARY_BLUE),
                     border=make_border())
        row += 1
        
        # Deliverables
        for deliverable in phase.get("deliverables", []):
            priority = deliverable.get("priority", "")
            priority_fill = get_priority_fill(priority)
            
            # Priority badge
            set_cell(ws, row, 1, priority,
                     font=make_font(BODY_TEXT, bold=True, size=11),
                     fill=priority_fill,
                     border=make_border(),
                     alignment=Alignment(horizontal="center"))
            
            # Deliverable name
            set_cell(ws, row, 2, deliverable.get("name", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            # Type
            set_cell(ws, row, 3, deliverable.get("type", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Why It Matters
            set_cell(ws, row, 4, deliverable.get("why_it_matters", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border(),
                     alignment=Alignment(wrap_text=True))
            
            # Dependencies
            set_cell(ws, row, 5, deliverable.get("dependencies", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            # Success Metric (orange styling)
            set_cell(ws, row, 6, deliverable.get("success_metric", ""),
                     font=make_font(ACCENT_ORANGE, bold=True, size=11),
                     border=make_border())
            
            # Owner
            set_cell(ws, row, 7, deliverable.get("owner", ""),
                     font=make_font(BODY_TEXT, size=11),
                     border=make_border())
            
            row += 1
        
        row += 2
    
    return ws


# =============================================================================
# MAIN BUILDER
# =============================================================================

def build_content_strategy(data, output_path):
    """Build the complete content strategy spreadsheet."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Build all sheets
    build_strategy_overview(wb, data)
    build_customer_profiles(wb, data)
    build_content_hub(wb, data)
    build_interactive_tools(wb, data)
    build_ebooks_guides(wb, data)
    build_roadmap(wb, data)
    
    # Save workbook
    wb.save(output_path)
    print(f"Content strategy spreadsheet saved to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Build content strategy spreadsheet")
    parser.add_argument("--data", required=True, help="Path to JSON data file")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    args = parser.parse_args()
    
    # Load data
    with open(args.data, "r") as f:
        data = json.load(f)
    
    # Build spreadsheet
    build_content_strategy(data, args.output)


if __name__ == "__main__":
    main()
